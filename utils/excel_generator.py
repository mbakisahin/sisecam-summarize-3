import pandas as pd
from openpyxl.comments import Comment
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import textwrap


class ExcelReportGenerator:
    """
    A class to handle the creation of Excel reports with comparison data, including hyperlinks, alternating row colors,
    and specific column formatting for the 'E' column and others.
    """

    def __init__(self):
        # Define gray and white fills for alternating row colors
        self.gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        self.white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        self.header_fill = PatternFill(start_color="1E90FF", end_color="1E90FF", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=12, name='Arial')
        self.thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                  bottom=Side(style='thin'))
        self.link_font = Font(color="0000FF", underline="single", name='Calibri', size=11)

    def wrap_text(self, text, width=100):
        """
        Wraps text after a certain width without splitting words.

        :param text: The text to wrap.
        :param width: The maximum number of characters per line.
        :return: Wrapped text with new lines, preserving whole words.
        """
        return textwrap.fill(text, width=width)

    def apply_header_style(self, worksheet, col_count):
        # Başlıkları biçimlendiren fonksiyon
        for col in range(1, col_count + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = self.header_font
            cell.border = self.thin_border

    def format_data_columns(self, worksheet, max_row):
        # Alternatif hücre renklendirmesi ve kenarlık işlemleri
        for col_idx in range(1, 6):
            fill_color = self.white_fill if col_idx % 2 == 0 else self.gray_fill
            for row_idx in range(2, max_row + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.fill = fill_color
                cell.border = self.thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def adjust_column_width(self, worksheet, start_idx, end_idx):
        # Sütun genişliğini metin uzunluğuna göre ayarlar
        for idx in range(start_idx, end_idx):
            col_letter = get_column_letter(idx)
            max_length = 0
            for cell in worksheet[col_letter]:
                if cell.value:
                    wrapped_text = self.wrap_text(str(cell.value))
                    max_length = max(max_length, max(len(line) for line in wrapped_text.split('\n')))
            worksheet.column_dimensions[col_letter].width = max_length

    def adjust_first_four_column_width(self, worksheet):
        # İlk 4 sütunun genişliğini sabit olarak ayarlayın
        worksheet.column_dimensions['A'].width = 20  # Direktörlük
        worksheet.column_dimensions['B'].width = 20  # Keyword
        worksheet.column_dimensions['C'].width = 20  # Date
        worksheet.column_dimensions['D'].width = 20  # Kaynak (Original Document linki)
        worksheet.column_dimensions['E'].width = 20  # Kaynak (Original Document linki)


    def add_neighbor_links_and_comparisons_with_comments(self, worksheet, neighbor_urls, individual_comparisons):
        # Komşu PDF linklerini ve karşılaştırma metinlerini ekler ve yorumları ekler
        for idx, (neighbor_url, comparison_text) in enumerate(zip(neighbor_urls, individual_comparisons), 1):
            worksheet.merge_cells(start_row=1, start_column=5 + 2 * idx - 1, end_row=1, end_column=5 + 2 * idx)
            merged_header_cell = worksheet.cell(row=1, column=5 + 2 * idx - 1)
            merged_header_cell.value = f"Benzer Doküman {idx}"
            merged_header_cell.fill = self.header_fill
            merged_header_cell.font = self.header_font
            merged_header_cell.alignment = Alignment(horizontal='center', vertical='center')
            merged_header_cell.border = self.thin_border

            fill_color = self.gray_fill if idx % 2 == 0 else self.white_fill

            # Link hücresini ekleyin
            link_cell = worksheet.cell(row=2, column=5 + 2 * idx - 1)
            link_cell.value = 'Link'
            link_cell.hyperlink = neighbor_url
            link_cell.font = self.link_font
            link_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            link_cell.border = self.thin_border
            link_cell.fill = fill_color

            # Karşılaştırma metnini ekleyin ve yorumu hücreye ekleyin
            comparison_cell = worksheet.cell(row=2, column=5 + 2 * idx)
            comparison_cell.value = "..."
            comparison_cell.font = Font(name='Arial', size=11)
            comparison_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            comparison_cell.border = self.thin_border
            comparison_cell.fill = fill_color

            # Yorum ekleme
            comment = Comment(self.wrap_text(comparison_text), "Comparison")
            comparison_cell.comment = comment

            # Sütun genişliklerini ayarlama
            worksheet.column_dimensions[get_column_letter(5 + 2 * idx - 1)].width = 15  # Link sütunu genişliği
            worksheet.column_dimensions[get_column_letter(5 + 2 * idx)].width = 10  # "..." sütunu genişliği

    def create_excel(self, metadata, file_name='comparison_report.xlsx'):
        # Verileri hazırlayın
        data = {
            'İlgili Direktörlük': ['Çevre'],
            'Keyword': [metadata.get('keyword', 'N/A')],
            'Date': [metadata.get('date', 'N/A')],
            'Kaynak': [''],
            'Key Differences': ['...']  # Sadece "..." gösterilecek
        }

        df = pd.DataFrame(data)
        neighbor_urls = metadata.get('neighbor_urls', [])
        individual_comparisons = metadata.get('individual_comparisons', [])

        for idx, (neighbor_url, comparison_text) in enumerate(zip(neighbor_urls, individual_comparisons), 1):
            comparison_column = f'Benzer Doküman {idx}'
            df[comparison_column] = [f"{self.wrap_text(comparison_text)}\n\nLink"]

        with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, startrow=0)
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']

            # Başlık biçimlendirme
            self.apply_header_style(worksheet, len(df.columns))

            # "Kaynak" başlığını ekleyin ve Original Document linkini tıklanabilir hale getirin
            worksheet.cell(row=1, column=4).value = "Kaynak"
            link_cell = worksheet.cell(row=2, column=4)
            original_url = metadata.get('url', '#')
            link_cell.value = 'Original Document'
            link_cell.hyperlink = original_url
            link_cell.font = self.link_font
            link_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            link_cell.border = self.thin_border

            # "Key Differences" için yorum ekleyin
            key_diff_cell = worksheet.cell(row=2, column=5)  # "Key Differences" hücresi
            key_diff_cell.value = "..."
            key_diff_cell.font = Font(name='Arial', size=11)
            key_diff_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            key_diff_cell.border = self.thin_border

            # "Key Differences" yorum ekleme
            key_diff_comment = Comment(self.wrap_text(metadata.get('combined_comparison', 'N/A')), "Key Differences")
            key_diff_cell.comment = key_diff_comment

            # Veri sütunlarını biçimlendirme (alternatif renklendirme ve kenarlık)
            self.format_data_columns(worksheet, worksheet.max_row)

            # Komşu linkleri ve karşılaştırma metinlerini ekleyin, metinleri yorum olarak ekleyin
            self.add_neighbor_links_and_comparisons_with_comments(worksheet, neighbor_urls, individual_comparisons)

            # İlk 4 sütunun genişliğini ayarla
            self.adjust_first_four_column_width(worksheet)

            # # Sütun genişliklerini ayarlama (diğer sütunlar için)
            # self.adjust_column_width(worksheet, 5, 6)
